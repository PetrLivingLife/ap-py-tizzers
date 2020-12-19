from openpyxl import load_workbook
from pathlib import Path


class Insert(object):
    def __init__(self):
        print("Inserting formulas into Priority and Step Number columns.")
        self._dir = r"\\to2.to2cz.cz\FS\SOC\05_Testing_all\05_E2E\02 Test analyza\05 Vystupni revize"
        step_number = '=IF(INDIRECT(ADDRESS(ROW(); 1))<>"";"Precondition.";IF(INDIRECT(ADDRESS(ROW()-1; COLUMN()))="Precondition.";1;INDIRECT(ADDRESS(ROW()-1; COLUMN()))+1))'.replace(";", ",")  # replace for excel storing ; as ,
        priority = '=IF(INDIRECT(ADDRESS(ROW(); 1))<>"";VLOOKUP(INDIRECT(ADDRESS(ROW(); 4));\'Test Cases\'!T:W;3;FALSE);INDIRECT(ADDRESS(ROW()-1; COLUMN())))'.replace(";", ",")
        for folder in Path(self._dir).iterdir():
            if folder.is_dir() and folder.name.startswith("BUS_"):
                for file in Path(folder).iterdir():
                    if (file.is_file() and (file.match("*\\TA_E2E_31_*ProvisioningOnly*.xls?") or file.match(
                            "*\\TA_E2E_31_*BillingOnly*.xls?"))):
                        print(file)
                        filename = str(file)
                        wb = load_workbook(filename=filename)
                        sh = wb["Test Scripts"]
                        for row in sh.iter_rows(min_row=2):
                            row[5].value = priority
                            row[7].value = step_number

                        wb.save(filename=filename)
                        wb.close()


if __name__ == '__main__':
    a = Insert()
